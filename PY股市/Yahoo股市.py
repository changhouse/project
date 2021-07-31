from bs4 import BeautifulSoup
import requests
import openpyxl
from openpyxl.styles import Font


class Stock:
    def __init__(self, *stock_numbers):
        self.stock_numbers = stock_numbers

    def scrape(self):

        result = list()

        for stock_number in self.stock_numbers:

            response = requests.get(
                "https://tw.stock.yahoo.com/q/q?s=" + stock_number)
            soup = BeautifulSoup(response.text.replace("加到投資組合", ""), "lxml")

            stock_date = soup.find(
                "font", {"class": "tt"}).getText().strip()[-9:]  # 資料日期

            tables = soup.find_all("table")[2]  # 取得網頁中第三個表格
            tds = tables.find_all("td")[0:11]  # 取得表格中1到10格

            result.append((stock_date,) +
                          tuple(td.getText().strip() for td in tds))
        return result


    def export(self, stocks):     #openpyxl模組(Module)的save()方法(Method)，傳入Excel檔案的名稱
        wb = openpyxl.Workbook()
        sheet = wb.create_sheet("Yahoo股市", 0)   # 新增一個新的sheet(工作表)，將它命名為「Yahoo股市」
        sheet.column_dimensions['B'].width = 15.0  # 調整列寬
        response = requests.get(
            "https://tw.stock.yahoo.com/q/q?s=2451")   # requests套件取得網頁的回應結果
        soup = BeautifulSoup(response.text, "lxml")    # 使用支援BeautifulSoup套件的lxml解析器來進行解析

        tables = soup.find_all("table")[2]   # find_all()方法(Method)，定位所有的表格(table)元素，並且只取得第三個表格(table)
        ths = tables.find_all("th")[0:11]    # 就能夠定位這個表格(table)元素下的前11個標題元素(th)
        titles = ("資料日期",) + tuple(th.getText() for th in ths)
        sheet.append(titles)

        for index, stock in enumerate(stocks):
            sheet.append(stock)

            if "△" in stock[6]:
                sheet.cell(row=index+2, column=7).font = Font(color='FF0000')  # 紅色
            elif "▽" in stock[6]:
                sheet.cell(row=index+2, column=7).font = Font(color='00A600')  # 綠色

        wb.save("yahoostock.xlsx")

stock = Stock('2451', '2454', '2369','2330','0050')  # 建立Stock物件

stock.export(stock.scrape())  # 將爬取的股票當日行情資料匯出成Excel檔案


