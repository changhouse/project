from bs4 import BeautifulSoup
import requests as req
import openpyxl
from openpyxl.styles import Font


class Inform:
    def __init__(self, *codename):
        self.codename = codename

    def scrape(self):
        result = list()
        for codename in self.codename:
            res = req.get(
                "https://tw.stock.yahoo.com/q/q?s=" + codename)
            bs = BeautifulSoup(res.text.replace("加到投資組合", ""), "lxml")
            date = bs.find(
                "font", {"class": "tt"}).getText().strip()[-9:]  # 日期
            tables = bs.find_all("table")[2]  # 取得網頁中第三個表格
            tds = tables.find_all("td")[0:11]  # 取得1 to 10格
            result.append((date,) +
                          tuple(td.getText().strip() for td in tds))
        return result


    def construct(self, stocks):     #openpyxl模組(Module)的save()方法(Method)，傳入Excel檔案的名稱
        wb = openpyxl.Workbook()
        sheet = wb.create_sheet("當日股市", 0)   # 新增一個新的sheet(工作表)，將它命名為「當日股市」
        sheet.column_dimensions['B'].width = 15.0  # 調整列寬
        res = req.get(
            "https://tw.stock.yahoo.com/q/q?s=2451")   # requests套件取得網頁的回應結果
        bs = BeautifulSoup(res.text, "lxml")    # 使用支援BeautifulSoup套件的lxml解析器來進行解析
        tables = bs.find_all("table")[2]   # find_all()方法(Method)，定位所有的表格(table)元素，並且只取得第三個表格(table)
        ths = tables.find_all("th")[0:11]    # 就能夠定位這個表格(table)元素下的前11個標題元素(th)
        titles = ("資料日期",) + tuple(th.getText() for th in ths)
        sheet.append(titles)
        for index, stock in enumerate(stocks):
            sheet.append(stock)

            if "△" in stock[6]:
                sheet.cell(row=index+2, column=7).font = Font(color='FF0000')  # 紅色
            if "▽" in stock[6]:
                sheet.cell(row=index+2, column=7).font = Font(color='00A600')  # 綠色

        wb.save("shareprice.xlsx")

stock = Inform('2303', '2344','2330','6214','2480','2454','8150','2002','0050')  

stock.construct(stock.scrape())  # 將爬取的股票當日行情資料匯出成Excel檔案


